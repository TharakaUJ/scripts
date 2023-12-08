from openpyxl import load_workbook
wbreadonly = load_workbook(filename = 'exprt1.xlsx', data_only=True)
wbwrite = load_workbook(filename= 'tmpfiles.org_dl_1947734_testacc.xlsx', data_only=True)
sheetWrite = wbwrite.sheetnames[0]
sheetRead = wbreadonly.sheetnames[0]

wsheetwrite = wbwrite[sheetWrite]
wsheetRead = wbreadonly[sheetRead]
NoRowsWrite = len(wsheetwrite['E'])
NoRowsRead = len(wsheetRead['A'])

for i in range(2, NoRowsWrite+1):
    string = wsheetwrite.cell(row=i, column=5).value
    wsheetwrite.cell(row=i, column=7).value = "-"
    wsheetwrite.cell(row=i, column=8).value = "-"

    for j in range(2, NoRowsRead):
        if wsheetRead.cell(row=j, column=1).value == string:
            PA_Read = int(wsheetRead.cell(row=j, column=3).value)
            DA_Read = int(wsheetRead.cell(row=j, column=4).value)
            wsheetwrite.cell(row=i, column=7).value = PA_Read
            wsheetwrite.cell(row=i, column=8).value = DA_Read
wbwrite.save('new1.xlsx')

        
